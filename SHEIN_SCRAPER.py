import os
import re
import sys
import threading
from datetime import date
from io import BytesIO

import requests
from PIL import Image
from PyQt5 import QtCore
from PyQt5.QtCore import QMutex, QMutexLocker
from PyQt5.QtWidgets import QMainWindow, QApplication, QVBoxLayout, QWidget, QTextEdit, QLabel, QProgressBar
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment, numbers
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

mutex = QMutex()


def add_data_to_excel_file(item_name, item_link, item_image, item_price, item_sizes):
    # $$$
    shipping = 900
    exchange_rate = 220

    us_price = float(item_price.replace("$", ''))

    gyd_price = us_price * exchange_rate + shipping

    filename = str(date.today()) + ".xlsx"

    try:
        workbook = load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active

    if not sheet["A1"].value:
        sheet["A1"] = "Product"
        sheet["B1"] = "Product Image"
        sheet["C1"] = "US Price"
        sheet["D1"] = "GYD Price + Shipping"
        sheet["E1"] = "Sizes"
        sheet["F1"] = "Page Price"
        sheet["G1"] = "Profit"

    item_hyperlink = '=HYPERLINK("{link}", "{name}")'.format(link=item_link, name=item_name)
    image_path = '=IMAGE("{image_url}")'.format(image_url=item_image)

    sheet.append([item_hyperlink, image_path, item_price, gyd_price, item_sizes, '', ''])

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        sheet.column_dimensions[col].width = 28

    header_font = Font(color="FFFFFF", bold=True)
    us_fill = PatternFill(fill_type="solid", fgColor="878787")
    gyd_fill = PatternFill(fill_type="solid", fgColor="004700")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in sheet["C"]:
        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        cell.font = header_font
        cell.fill = us_fill

    for cell in sheet["D"]:
        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        cell.font = header_font
        cell.fill = gyd_fill

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            sheet.row_dimensions[cell.row].height = 200

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    sheet.row_dimensions[1].height = 20

    workbook.save(filename)


class UpdateProgressSignal(QtCore.QObject):
    update_progress = QtCore.pyqtSignal(int)


class UpdateOutputSignal(QtCore.QObject):
    update_output = QtCore.pyqtSignal(str)


class SheinScraper(QMainWindow):
    def __init__(self):
        super().__init__()

        # open driver
        self.driver = None
        self.redirected_url = None
        self.chrome_options = Options()
        self.chrome_options.add_argument("--headless")
        self.chrome_options.add_argument("--disable-gpu")
        # self.driver = webdriver.Chrome(options=self.chrome_options)

        self.update_progress_signal = UpdateProgressSignal()
        self.update_progress_signal.update_progress.connect(self.update_progress_bar)

        self.update_output_signal = UpdateOutputSignal()
        self.update_output_signal.update_output.connect(self.update_output_window)

        self.sizes_str = None
        self.root_folder = None
        self.item_url = None
        self.folder_name = None

        self.setWindowTitle("Drag and Drop")
        self.setFixedSize(500, 400)
        self.setAcceptDrops(True)

        self.output_window = QTextEdit()
        self.output_window.setReadOnly(True)
        self.output_window.setFixedHeight(250)

        self.label = QLabel("Drag file here")
        self.label.setAlignment(QtCore.Qt.AlignCenter)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setAlignment(QtCore.Qt.AlignCenter)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.output_window)
        central_widget.setLayout(layout)

    def update_output_window(self, text):
        self.output_window.append(text)

    def update_progress_bar(self, value):
        self.progress_bar.setValue(value)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def scroll_output(self):
        scrollbar = self.output_window.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def dropEvent(self, event):
        self.root_folder = str(date.today())
        os.makedirs(self.root_folder, exist_ok=True)

        files = [u.toLocalFile() for u in event.mimeData().urls()]
        for file_path in files:
            self.update_output_signal.update_output.emit(f"------->  Grabbing Product Details from: {file_path}")
            self.scroll_output()

            # start a thread for these items
            threading.Thread(target=self.get_item_urls, args=(file_path,)).start()

    def save_image_as_jpg(self, url, filepath):
        response = requests.get(url)
        with Image.open(BytesIO(response.content)) as img:
            img = img.convert('RGB')
            img.save(filepath, 'JPEG')

    def get_redirected_url(self, url):

        # for mobile share links
        if 'api' in url:
            self.driver.get(url)
            WebDriverWait(self.driver, 1).until(EC.url_changes(url))

            self.redirected_url = self.driver.current_url
        else:
            self.redirected_url = url

        return self.redirected_url

    def get_item_data(self, url):
        url = self.get_redirected_url(url)
        self.item_url = url

        sizes = []
        picture_urls = []
        price = None

        self.driver.get(url)
        self.driver.implicitly_wait(5)
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')

        item_name = soup.title.text.strip()[:50]
        item_name = re.sub(r'[<>:"/\\|?*]', '_', item_name)
        item_name = item_name.replace(' ', '_')

        self.update_output_signal.update_output.emit(f"------->  Processing: {item_name}")
        self.scroll_output()

        # image scraping
        picture_url_tag = soup.find_all('div', class_='product-intro-zoom__item')
        picture_thumbnails = soup.find_all('div', class_='product-intro__thumbs-item')

        for element in picture_url_tag:
            img_tag = element.find('img', class_='lazyload crop-image-container__img')
            if img_tag:
                url = img_tag['data-src']
                if url.startswith("//"):
                    url = "https:" + url
                picture_urls.append(url)

        for element in picture_thumbnails:
            img_tag = element.find('img', class_='fsp-element crop-image-container__img')
            if img_tag:
                url = img_tag['src']
                if url.startswith("//"):
                    url = "https:" + url
                picture_urls.append(url)

        # find the sizes
        size_elements = soup.find_all('div', class_='product-intro__size-choose')
        for size_element in size_elements:
            size_items = size_element.find_all('div', class_='product-intro__size-radio')
            for size_item in size_items:
                size = size_item.get('data-attr_value_name')
                if size:
                    sizes.append(size)

        # first price it finds, could be sale price could be the non sale price
        elements = soup.find_all('div', class_='from original')
        for element in elements:
            if "$" in element.text:
                price = element.text
                break

        # make folders
        self.folder_name = item_name
        os.makedirs(os.path.join(self.root_folder, self.folder_name), exist_ok=True)

        self.sizes_str = "_".join(sizes)

        # save picutes exclude some overlays
        for i, picture_url in enumerate(picture_urls):
            image_size = int(requests.head(picture_url).headers['Content-Length'])
            # dont download overlays
            if image_size < 5 * 1024:
                continue

            filename = f"{item_name}_{price}_{self.sizes_str}_{i}.jpg"
            filepath = os.path.join(self.root_folder, self.folder_name, filename)

            self.save_image_as_jpg(picture_url, filepath)
            self.scroll_output()

        # add item data to file
        add_data_to_excel_file(item_name, self.item_url, picture_urls[1], price, self.sizes_str)

        self.update_output_signal.update_output.emit(f"------->  Done")
        self.scroll_output()

    def get_item_urls(self, file_name: str):

        # start the driver
        self.driver = webdriver.Chrome(options=self.chrome_options)
        self.progress_bar.setValue(0)

        with open(file_name, "r") as f:
            urls = f.readlines()

        total_urls = len([url.strip() for url in urls if url.strip()])
        self.progress_bar.setMaximum(total_urls)

        for idx, item_url in enumerate(urls):
            item_url = item_url.strip()
            self.scroll_output()
            if not item_url:
                continue

            try:
                self.get_item_data(item_url)
            except Exception as e:
                with QMutexLocker(mutex):
                    self.update_output_signal.update_output.emit(f"Error processing URL: {item_url}. Error: {str(e)}")

            finally:
                with QMutexLocker(mutex):
                    self.update_progress_signal.update_progress.emit(idx + 1)

        self.driver.quit()
        self.update_output_signal.update_output.emit(f"------->  Done: {file_name} \n")
        self.scroll_output()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ui = SheinScraper()
    ui.show()
    sys.exit(app.exec_())
